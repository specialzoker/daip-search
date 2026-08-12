"""대입통합검색기_데이터정리본.xlsx → main.csv / min.csv / jong.csv

데이터 업데이트 절차:
  1. 대입통합검색기_데이터정리본.xlsx 에서 데이터 수정
  2. python build_csv.py 실행
  3. git add data/ && git commit && git push
"""
import openpyxl, csv, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\user\search\대입통합검색기_데이터템플릿.xlsx'
DATA_DIR = r'C:\Users\user\search\data'
os.makedirs(DATA_DIR, exist_ok=True)

SHEETS = [
    ('메인',     'main.csv'),
    ('최저기준', 'min.csv'),
    ('전형방법', 'jong.csv'),  # 시트명 '종합전형' → '전형방법'으로 변경됨
]

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

for sheet_name, out_name in SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f'⚠ {sheet_name} 시트 없음')
        continue
    ws = wb[sheet_name]
    out_path = os.path.join(DATA_DIR, out_name)
    rows = []
    for row in ws.iter_rows(values_only=True):
        # 모든 값 None인 행은 건너뜀
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        rows.append([('' if v is None else str(v).strip()) for v in row])
    # 메인 시트: 중복된 '충원' 헤더를 직전 전형 기준으로 고유화 (교과1/충원 등)
    if sheet_name == '메인' and rows:
        hdr = rows[0]
        prefix = None
        for i, h in enumerate(hdr):
            if '/전형' in h:
                prefix = h.split('/')[0]      # 교과1, 교과2, 종합1 ...
            elif h == '정시백분위':
                prefix = '정시'
            elif h == '충원' and prefix:
                hdr[i] = prefix + '/충원'
        rows[0] = hdr
        # 권역 정규화 — 데이터를 검색기 6권역 필터에 맞춤
        KWON_MAP = {'서울권':'서울', '부울경권':'경상권', '대경권':'경상권', '호남권':'전라권'}
        kwon_idx = hdr.index('권역') if '권역' in hdr else 0
        for r in rows[1:]:
            if kwon_idx < len(r) and r[kwon_idx] in KWON_MAP:
                r[kwon_idx] = KWON_MAP[r[kwon_idx]]
        # 중복 모집단위 정리 — '학부' 접두사만 다른 중복은 완성도 높은 행만 남김
        # (단과대학 '○○대학' 접두는 보존 → 단과대별 별도 모집 유지)
        univ_i = hdr.index('대학명') if '대학명' in hdr else None
        dept_i = hdr.index('2026 수시 모집단위') if '2026 수시 모집단위' in hdr else None
        if univ_i is not None and dept_i is not None:
            cut_cols = [hdr.index(c) for c in
                        ['교과1/70%','교과2/70%','교과3/70%','종합1/70%','종합2/70%','정시백분위'] if c in hdr]
            def norm_dept(d):
                d = d.strip()
                m = re.match(r'^\S*학부\s+(.+)$', d)   # '○○학부 △△전공' → '△△전공'
                if m: d = m.group(1)
                d = d.replace(' ', '')
                for suf in ('학과','학부'):  # 학과↔학부 접미사 통일
                    if d.endswith(suf) and len(d) > len(suf) + 1:
                        return d[:-len(suf)]
                return d
            def score(r):
                return sum(1 for c in cut_cols if c < len(r) and r[c].strip())
            groups = {}
            for r in rows[1:]:
                groups.setdefault((r[univ_i].strip(), norm_dept(r[dept_i])), []).append(r)
            chosen = set()
            newrows = [hdr]
            for r in rows[1:]:
                key = (r[univ_i].strip(), norm_dept(r[dept_i]))
                g = groups[key]
                if len(g) == 1:
                    newrows.append(r); continue
                if key in chosen:
                    continue                         # 그룹 대표 이미 추가됨 → 중복 제거
                newrows.append(max(g, key=score))    # 완성도 최고 행만 (동점 시 첫째)
                chosen.add(key)
            removed = len(rows) - len(newrows)
            rows = newrows
            if removed:
                print(f'    └ 중복 모집단위 {removed}건 정리됨')

            # 정시 입결 머지 (data/jeongsi.csv — 어디가 정시 백분위)
            jpath = os.path.join(DATA_DIR, 'jeongsi.csv')
            if os.path.exists(jpath):
                import statistics
                with open(jpath, encoding='utf-8') as jf:
                    jrows = list(csv.DictReader(jf))
                jnu = lambda s: re.sub(r'\s','',str(s or '')).replace('대학교','대').replace('학교','')
                jnd = lambda s: re.sub(r'\s','',str(s or ''))
                SMAP = {'인문・사회계열':'인문','인문·사회계열':'인문','공학계열':'자연',
                        '자연과학계열':'자연','의학계열':'자연','예체능계열':'예체능'}
                def fnum(x):
                    try: return float(str(x).replace(',',''))
                    except: return None
                # 정확매칭: (대학,모집단위) → 대표행 (일반전형 우선, 백분위 최대)
                exact = {}
                for jr in jrows:
                    bp = fnum(jr.get('백분위'))
                    if bp is None: continue
                    k = (jnu(jr['대학명']), jnd(jr['모집단위']))
                    rank = (1 if '일반' in jr.get('전형','') else 0, bp)
                    if k not in exact or rank > exact[k][0]:
                        exact[k] = (rank, jr)
                # 계열폴백: (대학,계열) → 백분위 중앙값 (의학계열 제외, 일반전형만)
                pool = {}
                for jr in jrows:
                    if str(jr.get('계열','')).strip() == '의학계열': continue
                    if '일반' not in jr.get('전형',''): continue
                    g = SMAP.get(str(jr.get('계열','')).strip(), '')
                    bp = fnum(jr.get('백분위'))
                    if not g or bp is None: continue
                    pool.setdefault((jnu(jr['대학명']), g), []).append(bp)
                fb = {k: statistics.median(v) for k, v in pool.items()}
                pct_i  = hdr.index('정시백분위')  if '정시백분위'  in hdr else None
                chu_i  = hdr.index('정시/충원')   if '정시/충원'   in hdr else None
                cap_i  = hdr.index('모집인원_정시') if '모집인원_정시' in hdr else None
                comp_i = hdr.index('경쟁률_정시')  if '경쟁률_정시'  in hdr else None
                ser_i  = hdr.index('계열')        if '계열'        in hdr else None
                maxi = max(x for x in [pct_i,chu_i,cap_i,comp_i,ser_i] if x is not None)
                ce = cf = 0
                for r in rows[1:]:
                    while len(r) <= maxi: r.append('')
                    k = (jnu(r[univ_i]), jnd(r[dept_i]))
                    if k in exact:
                        jr = exact[k][1]
                        if pct_i  is not None: r[pct_i] = str(jr['백분위'])
                        if chu_i  is not None and str(jr.get('충원','')).strip():   r[chu_i]  = str(jr['충원'])
                        if cap_i  is not None and str(jr.get('모집인원','')).strip(): r[cap_i]  = str(jr['모집인원'])
                        if comp_i is not None and str(jr.get('경쟁률','')).strip():   r[comp_i] = str(jr['경쟁률'])
                        ce += 1
                    elif pct_i is not None and ser_i is not None and not r[pct_i].strip():
                        fk = (jnu(r[univ_i]), r[ser_i].strip())
                        if fk in fb:
                            r[pct_i] = str(round(fb[fk], 1)); cf += 1
                print(f'    └ 정시 머지: 정확 {ce}건, 계열폴백 {cf}건')

            # 나비 데이터 머지 (data/navi.csv) — 5등급 컷 / 2027 모집단위 변경 / 정시 보완
            npath = os.path.join(DATA_DIR, 'navi.csv')
            if os.path.exists(npath):
                nnu = lambda s: re.sub(r'\s','',str(s or '')).replace('대학교','대').replace('학교','')
                def nnd(s):   # 학과↔학부 접미사 통일 (전자공학과 ↔ 전자공학부)
                    d = re.sub(r'\s','',str(s or ''))
                    for suf in ('학과','학부'):
                        if d.endswith(suf) and len(d) > len(suf) + 1:
                            return d[:-len(suf)]
                    return d
                with open(npath, encoding='utf-8') as nf:
                    nmap = {(nnu(r['대학명']), nnd(r['모집단위'])): r for r in csv.DictReader(nf)}
                cut5 = [('교과1_50_5','교과1/50%(5등급)'),('교과1_70_5','교과1/70%(5등급)'),
                        ('교과2_50_5','교과2/50%(5등급)'),('교과2_70_5','교과2/70%(5등급)'),
                        ('교과3_50_5','교과3/50%(5등급)'),('교과3_70_5','교과3/70%(5등급)'),
                        ('종합1_50_5','종합1/50%(5등급)'),('종합1_70_5','종합1/70%(5등급)'),
                        ('종합2_50_5','종합2/50%(5등급)'),('종합2_70_5','종합2/70%(5등급)')]
                for _, dst in cut5:
                    if dst not in hdr: hdr.append(dst)
                if '2027모집단위' not in hdr: hdr.append('2027모집단위')
                if '26정시코드' not in hdr: hdr.append('26정시코드')
                colidx = {dst: hdr.index(dst) for _, dst in cut5}
                d27_i  = hdr.index('2027모집단위')
                code_i = hdr.index('26정시코드')
                pct_i2 = hdr.index('정시백분위')       if '정시백분위'       in hdr else None
                eng_i2 = hdr.index('영어한국사')       if '영어한국사'       in hdr else None
                nc5 = nc27 = 0
                for r in rows[1:]:
                    while len(r) < len(hdr): r.append('')
                    nr = nmap.get((nnu(r[univ_i]), nnd(r[dept_i])))
                    if not nr: continue
                    got5 = False
                    for src, dst in cut5:
                        if nr.get(src,''):
                            r[colidx[dst]] = nr[src]; got5 = True
                    if got5: nc5 += 1
                    if nr.get('모집단위2027','') and nr['모집단위2027'] != nr['모집단위']:
                        r[d27_i] = nr['모집단위2027']; nc27 += 1
                    if nr.get('정시코드','') and not r[code_i].strip():
                        r[code_i] = nr['정시코드']
                    if pct_i2 is not None and not r[pct_i2].strip() and nr.get('정시백분위',''):
                        r[pct_i2] = nr['정시백분위']
                    if eng_i2 is not None and not r[eng_i2].strip() and nr.get('정시영어한국사',''):
                        r[eng_i2] = nr['정시영어한국사']
                print(f'    └ 나비 머지: 5등급컷 {nc5}건, 2027변경 {nc27}건')

                # 나비 신규 학과 추가 (main 미매칭 + 일반대만, 특수목적대 제외)
                mkey = set((nnu(r[univ_i]), nnd(r[dept_i])) for r in rows[1:])
                SKIP_KW = ['사관','경찰대','신학','승가']
                SKIP_UNIV = {'광신대','수원가톨릭대'}
                is_special = lambda u: any(k in u for k in SKIP_KW) or u in SKIP_UNIV
                KWON_MAP2 = {'서울권':'서울','부울경권':'경상권','대경권':'경상권','호남권':'전라권','수도권':'경인권'}
                SERIES_MAP2 = {'의약학':'자연','무계열':'공통'}
                col_map = {
                    '권역':'권역','지역':'지역','세부지역':'세부지역','대학명':'대학명',
                    '2026 수시 모집단위':'모집단위','계열':'계열',
                    '교과1/전형':'교과1전형','교과1/50%':'교과1_50','교과1/70%':'교과1_70','교과1/50%(5등급)':'교과1_50_5','교과1/70%(5등급)':'교과1_70_5',
                    '교과2/전형':'교과2전형','교과2/50%':'교과2_50','교과2/70%':'교과2_70','교과2/50%(5등급)':'교과2_50_5','교과2/70%(5등급)':'교과2_70_5',
                    '교과3/전형':'교과3전형','교과3/50%':'교과3_50','교과3/70%':'교과3_70','교과3/50%(5등급)':'교과3_50_5','교과3/70%(5등급)':'교과3_70_5',
                    '종합1/전형':'종합1전형','종합1/50%':'종합1_50','종합1/70%':'종합1_70','종합1/50%(5등급)':'종합1_50_5','종합1/70%(5등급)':'종합1_70_5',
                    '종합2/전형':'종합2전형','종합2/50%':'종합2_50','종합2/70%':'종합2_70','종합2/50%(5등급)':'종합2_50_5','종합2/70%(5등급)':'종합2_70_5',
                    '정시백분위':'정시백분위','영어한국사':'정시영어한국사','2025 정시 반영영역':'정시반영영역',
                }
                added = 0
                for nr in nmap.values():
                    u = nr.get('대학명',''); d = nr.get('모집단위','')
                    if not u or not d or (nnu(u), nnd(d)) in mkey or is_special(u):
                        continue
                    newr = [''] * len(hdr)
                    for mcol, ncol in col_map.items():
                        if mcol in hdr and nr.get(ncol,''):
                            v = nr[ncol]
                            if mcol == '권역': v = KWON_MAP2.get(v, v)
                            elif mcol == '계열': v = SERIES_MAP2.get(v, v)
                            newr[hdr.index(mcol)] = v
                    if nr.get('모집단위2027','') and nr['모집단위2027'] != d and '2027모집단위' in hdr:
                        newr[hdr.index('2027모집단위')] = nr['모집단위2027']
                    if nr.get('정시코드','') and '26정시코드' in hdr:
                        newr[hdr.index('26정시코드')] = nr['정시코드']
                    rows.append(newr)
                    mkey.add((nnu(u), nnd(d)))
                    added += 1
                print(f'    └ 나비 신규 학과 추가: {added}건 (일반대만)')

            # 2027 수시요강 70%컷 머지 (data/cut2027.csv) — 빈 슬롯만 채움
            cpath = os.path.join(DATA_DIR, 'cut2027.csv')
            if os.path.exists(cpath):
                cnu = lambda s: re.sub(r'\s','',str(s or '')).replace('대학교','대').replace('학교','')
                cnu2 = lambda s: re.sub(r'\([^)]*\)','', cnu(s))   # 캠퍼스 표기 제거
                def _sufcut(d):
                    for suf in ('학과','학부','전공'):
                        if d.endswith(suf) and len(d) > len(suf) + 1: return d[:-len(suf)]
                    return d
                def cnd(s):
                    d = re.sub(r'\s','',str(s or ''))
                    d = re.sub(r'^\[[^\]]*\]','', d)               # '[유형1] ' 접두 제거
                    return _sufcut(d)
                def cnd3(s):                                        # 끝 괄호(계열) 제거: 간호학과(인문) → 간호
                    d = re.sub(r'\s','',str(s or ''))
                    d = re.sub(r'^\[[^\]]*\]','', d)
                    return _sufcut(re.sub(r'\([^)]*\)$','', d))
                with open(cpath, encoding='utf-8') as cf:
                    cuts_all = list(csv.DictReader(cf))
                if True:
                    cmap, cmap2, cmap3 = {}, {}, {}
                    for cr in cuts_all:
                        base = cr.get('기준','')
                        if '환산' in base: continue          # 환산점수는 등급이 아니므로 컷으로 못 씀
                        try: v = float(cr['컷70'])
                        except: continue
                        if not (0.5 <= v <= 9): continue
                        cmap.setdefault((cnu(cr['대학명']), cnd(cr['모집단위']), cr['유형']), []).append(cr)
                        cmap2.setdefault((cnu2(cr['대학명']), cnd(cr['모집단위']), cr['유형']), []).append(cr)
                        cmap3.setdefault((cnu2(cr['대학명']), cnd3(cr['모집단위']), cr['유형']), []).append(cr)
                # 전형명 정규화·유사도 (요강은 '학생부교과(추천형)' 형식)
                def ctn(s):
                    v = str(s or '')
                    m = re.search(r'\(([^)]*)\)\s*$', v)
                    if m and len(m.group(1)) >= 2: v = m.group(1)
                    v = re.sub(r'학생부종합|학생부교과|논술위주|논술|전형|형$', '', v)
                    return re.sub(r'[^0-9A-Za-z가-힣]', '', v)
                def tsim(a, b):
                    if not a or not b: return 0
                    if a == b: return 1
                    if a in b or b in a: return 0.9
                    sh, lo = (a, b) if len(a) <= len(b) else (b, a)
                    g = [sh[i:i+2] for i in range(len(sh)-1)] or [sh]
                    return sum(1 for x in g if x in lo)/len(g)
                slots = {'교과': [('교과1/전형','교과1/70%'),('교과2/전형','교과2/70%'),('교과3/전형','교과3/70%')],
                         '종합': [('종합1/전형','종합1/70%'),('종합2/전형','종합2/70%')]}
                fill_cut = fill_new = 0
                for r in rows[1:]:
                    for kind, pairs in slots.items():
                        items = (cmap.get((cnu(r[univ_i]), cnd(r[dept_i]), kind))
                                 or cmap2.get((cnu2(r[univ_i]), cnd(r[dept_i]), kind))
                                 or cmap3.get((cnu2(r[univ_i]), cnd3(r[dept_i]), kind)))
                        if not items: continue
                        idxs = [(hdr.index(a), hdr.index(b)) for a, b in pairs if a in hdr and b in hdr]
                        used = set()
                        for it in items:
                            nit = ctn(it['전형명'])
                            # 1) 같은 전형이 이미 있고 컷만 비었으면 컷 채움
                            hit = None
                            for ti, ci in idxs:
                                if (ti, ci) in used or not r[ti].strip(): continue
                                if tsim(nit, ctn(r[ti])) >= 0.5: hit = (ti, ci); break
                            if hit:
                                ti, ci = hit; used.add(hit)
                                if not r[ci].strip():
                                    r[ci] = it['컷70']; fill_cut += 1
                                continue
                            # 2) 없으면 완전히 빈 슬롯에 전형명+컷 추가
                            for ti, ci in idxs:
                                if (ti, ci) in used: continue
                                if not r[ti].strip() and not r[ci].strip():
                                    r[ti] = it['전형명']; r[ci] = it['컷70']
                                    used.add((ti, ci)); fill_new += 1
                                    break
                print(f'    └ 2027요강 70%컷 보완: 기존전형 컷채움 {fill_cut}건, 신규전형 {fill_new}건')

                # 광역(단과대/계열) 모집 컷 → 해당 계열 학과에 참고값으로 부여
                # 예: 중앙대 '[유형2] 자연과학대학' 1.59 → 중앙대 자연계열 학과들
                def guess_series(name):
                    v = re.sub(r'^\[[^\]]*\]\s*', '', str(name or ''))
                    if re.search(r'예술|체육|음악|미술|디자인|무용|연극|공연', v): return '예체능'
                    if re.search(r'자연|공과|공학|과학|IT|ICT|SW|AI|소프트|정보|의과|약학|간호|보건|생명|바이오|융합', v, re.I): return '자연'
                    if re.search(r'인문|사회|경영|경제|법|사범|교육|글로벌|관광|문화|미디어|상경|어문', v): return '인문'
                    return ''
                broad_re = re.compile(r'(대학|계열)$')
                broad = {}
                for cr in cuts_all:
                    nm = re.sub(r'^\[[^\]]*\]\s*', '', cr['모집단위'].strip())
                    if not broad_re.search(nm): continue
                    ser = guess_series(nm)
                    if not ser: continue
                    try: v = float(cr['컷70'])
                    except: continue
                    if not (0.5 <= v <= 9): continue
                    broad.setdefault((cnu2(cr['대학명']), ser, cr['유형']), []).append((v, nm, cr['전형명']))
                for col in ('교과광역/70%','교과광역/출처','종합광역/70%','종합광역/출처'):
                    if col not in hdr: hdr.append(col)
                bg70, bgsrc = hdr.index('교과광역/70%'), hdr.index('교과광역/출처')
                bj70, bjsrc = hdr.index('종합광역/70%'), hdr.index('종합광역/출처')
                ser_i2 = hdr.index('계열') if '계열' in hdr else None
                bfill = 0
                for r in rows[1:]:
                    while len(r) < len(hdr): r.append('')
                    if ser_i2 is None: break
                    for kind, (vi, si), cols in (('교과', (bg70, bgsrc), ['교과1/70%','교과2/70%','교과3/70%']),
                                                 ('종합', (bj70, bjsrc), ['종합1/70%','종합2/70%'])):
                        if any(r[hdr.index(c)].strip() for c in cols if c in hdr): continue   # 학과 자체 컷 있으면 skip
                        cand = broad.get((cnu2(r[univ_i]), r[ser_i2].strip(), kind))
                        if not cand: continue
                        v, nm, tname = min(cand, key=lambda x: x[0])   # 여러 개면 가장 높은 성적대(작은 값)
                        r[vi] = f'{v}'; r[si] = f'{nm} · {tname}'
                        bfill += 1
                print(f'    └ 광역(단과대/계열) 참고컷 부여: {bfill}건')

                # 환산점수 발표 대학(한국외대 등) — 등급이 아니라 참고 표시용으로 저장
                conv = {}
                for cr in cuts_all:
                    if '환산' not in cr.get('기준',''): continue
                    v = cr.get('컷70','').strip()
                    if not v: continue
                    for key in ((cnu(cr['대학명']), cnd(cr['모집단위']), cr['유형']),
                                (cnu2(cr['대학명']), cnd3(cr['모집단위']), cr['유형'])):
                        conv.setdefault(key, (v, cr['기준'], cr['전형명']))
                for col in ('교과환산/70%','종합환산/70%'):
                    if col not in hdr: hdr.append(col)
                cg, cj = hdr.index('교과환산/70%'), hdr.index('종합환산/70%')
                cfill = 0
                for r in rows[1:]:
                    while len(r) < len(hdr): r.append('')
                    for kind, ci in (('교과', cg), ('종합', cj)):
                        hit = (conv.get((cnu(r[univ_i]), cnd(r[dept_i]), kind))
                               or conv.get((cnu2(r[univ_i]), cnd3(r[dept_i]), kind)))
                        if hit and not r[ci].strip():
                            r[ci] = f'{hit[0]} ({hit[2]})'; cfill += 1
                print(f'    └ 환산점수 발표 대학 참고값: {cfill}건')
    with open(out_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerows(rows)
    print(f'  {out_name}: {len(rows)-1:,} 데이터 행 ({os.path.getsize(out_path)/1024:.1f} KB)')

print('\n완료. 다음 명령으로 배포:')
print('  git add data/ && git commit -m "데이터 업데이트" && git push')
